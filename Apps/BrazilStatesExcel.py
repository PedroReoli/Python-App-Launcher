import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import os
import json
import threading
from functools import partial

# Cores personalizadas
PRIMARY_COLOR = "#56239f"
SECONDARY_COLOR = "#ff7200"
WHITE_COLOR = "#ffffff"
BLACK_COLOR = "#000000"
LIGHT_GRAY = "#f8f8f8"
MEDIUM_GRAY = "#e0e0e0"
DARK_GRAY = "#9e9e9e"

class EstadosApp:
    def __init__(self, root):
        self.root = root
        
        # Configuração da janela principal
        self.root.title("Gerenciador de Combinações de Estados")
        self.root.geometry("1200x800")
        self.root.minsize(800, 600)  # Tamanho mínimo para garantir usabilidade
        self.root.configure(bg=WHITE_COLOR)
        
        # Configurar o grid para responsividade
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(1, weight=1)  # Conteúdo principal
        
        # Lista de estados brasileiros
        self.estados = ["AC", "AL", "AM", "AP", "BA", "CE", "DF", "ES", "GO", "MA", 
                        "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", 
                        "RS", "RO", "RR", "SC", "SP", "SE", "TO"]
        
        # Dicionário para armazenar os valores
        self.valores = {}
        for origem in self.estados:
            for destino in self.estados:
                self.valores[f"{origem}-{destino}"] = ""
        
        # Variáveis para controle de edição
        self.current_edit = None
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self.filter_table)
        
        # Configurar estilo para ttk widgets
        self.setup_styles()
        
        # Criar widgets
        self.create_widgets()
        
    def setup_styles(self):
        # Configurar estilos para ttk widgets
        style = ttk.Style()
        
        # Estilo para Combobox
        style.configure("TCombobox", 
                        fieldbackground=WHITE_COLOR, 
                        background=WHITE_COLOR,
                        foreground=BLACK_COLOR,
                        arrowcolor=PRIMARY_COLOR)
        
        # Estilo para Scrollbar
        style.configure("TScrollbar", 
                        background=PRIMARY_COLOR, 
                        troughcolor=LIGHT_GRAY,
                        arrowcolor=WHITE_COLOR)
        
        # Estilo para Entry
        style.configure("TEntry", 
                        fieldbackground=WHITE_COLOR,
                        foreground=BLACK_COLOR)
        
        # Estilo para Frame
        style.configure("TFrame", background=WHITE_COLOR)
        
    def create_widgets(self):
        # Frame principal com padding
        main_container = tk.Frame(self.root, bg=WHITE_COLOR, padx=15, pady=15)
        main_container.grid(row=0, column=0, sticky="nsew")
        main_container.grid_columnconfigure(0, weight=1)
        main_container.grid_rowconfigure(1, weight=1)
        
        # ===== CABEÇALHO =====
        header_frame = tk.Frame(main_container, bg=WHITE_COLOR)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 15))
        header_frame.grid_columnconfigure(0, weight=1)
        
        # Título e subtítulo
        title_frame = tk.Frame(header_frame, bg=WHITE_COLOR)
        title_frame.grid(row=0, column=0, sticky="w")
        
        title_label = tk.Label(
            title_frame, 
            text="Gerenciador de Combinações de Estados", 
            font=("Arial", 22, "bold"),
            fg=PRIMARY_COLOR,
            bg=WHITE_COLOR
        )
        title_label.pack(anchor="w")
        
        subtitle_label = tk.Label(
            title_frame, 
            text="Preencha os valores para cada combinação de estados", 
            font=("Arial", 12),
            fg=DARK_GRAY,
            bg=WHITE_COLOR
        )
        subtitle_label.pack(anchor="w", pady=(0, 10))
        
        # ===== BARRA DE FERRAMENTAS =====
        toolbar_frame = tk.Frame(header_frame, bg=WHITE_COLOR, pady=10)
        toolbar_frame.grid(row=1, column=0, sticky="ew")
        toolbar_frame.grid_columnconfigure(7, weight=1)  # Empurra a barra de pesquisa para a direita
        
        # Estilo para botões
        button_style = {
            "font": ("Arial", 10),
            "borderwidth": 0,
            "padx": 15,
            "pady": 8,
            "cursor": "hand2",
            "relief": tk.RAISED
        }
        
        # Botões de arquivo
        file_label = tk.Label(
            toolbar_frame, 
            text="Arquivo:", 
            font=("Arial", 10, "bold"),
            fg=PRIMARY_COLOR,
            bg=WHITE_COLOR
        )
        file_label.grid(row=0, column=0, padx=(0, 5))
        
        save_button = tk.Button(
            toolbar_frame, 
            text="Salvar", 
            command=self.save_data,
            bg=PRIMARY_COLOR,
            fg=WHITE_COLOR,
            activebackground=SECONDARY_COLOR,
            activeforeground=WHITE_COLOR,
            **button_style
        )
        save_button.grid(row=0, column=1, padx=2)
        
        load_button = tk.Button(
            toolbar_frame, 
            text="Carregar", 
            command=self.load_data,
            bg=PRIMARY_COLOR,
            fg=WHITE_COLOR,
            activebackground=SECONDARY_COLOR,
            activeforeground=WHITE_COLOR,
            **button_style
        )
        load_button.grid(row=0, column=2, padx=2)
        
        # Separador
        separator1 = tk.Frame(toolbar_frame, width=1, height=25, bg=MEDIUM_GRAY)
        separator1.grid(row=0, column=3, padx=10)
        
        # Botões de exportação
        export_label = tk.Label(
            toolbar_frame, 
            text="Exportar:", 
            font=("Arial", 10, "bold"),
            fg=PRIMARY_COLOR,
            bg=WHITE_COLOR
        )
        export_label.grid(row=0, column=4, padx=(0, 5))
        
        export_csv_button = tk.Button(
            toolbar_frame, 
            text="CSV", 
            command=self.export_to_csv,
            bg=PRIMARY_COLOR,
            fg=WHITE_COLOR,
            activebackground=SECONDARY_COLOR,
            activeforeground=WHITE_COLOR,
            **button_style
        )
        export_csv_button.grid(row=0, column=5, padx=2)
        
        export_excel_button = tk.Button(
            toolbar_frame, 
            text="Excel", 
            command=self.export_to_excel,
            bg=PRIMARY_COLOR,
            fg=WHITE_COLOR,
            activebackground=SECONDARY_COLOR,
            activeforeground=WHITE_COLOR,
            **button_style
        )
        export_excel_button.grid(row=0, column=6, padx=2)
        
        # Separador
        separator2 = tk.Frame(toolbar_frame, width=1, height=25, bg=MEDIUM_GRAY)
        separator2.grid(row=0, column=7, padx=10)
        
        # Botões de importação
        import_label = tk.Label(
            toolbar_frame, 
            text="Importar:", 
            font=("Arial", 10, "bold"),
            fg=PRIMARY_COLOR,
            bg=WHITE_COLOR
        )
        import_label.grid(row=0, column=8, padx=(0, 5))
        
        import_csv_button = tk.Button(
            toolbar_frame, 
            text="CSV", 
            command=self.import_from_csv,
            bg=SECONDARY_COLOR,
            fg=WHITE_COLOR,
            activebackground=PRIMARY_COLOR,
            activeforeground=WHITE_COLOR,
            **button_style
        )
        import_csv_button.grid(row=0, column=9, padx=2)
        
        import_excel_button = tk.Button(
            toolbar_frame, 
            text="Excel", 
            command=self.import_from_excel,
            bg=SECONDARY_COLOR,
            fg=WHITE_COLOR,
            activebackground=PRIMARY_COLOR,
            activeforeground=WHITE_COLOR,
            **button_style
        )
        import_excel_button.grid(row=0, column=10, padx=2)
        
        # Barra de pesquisa
        search_frame = tk.Frame(toolbar_frame, bg=WHITE_COLOR)
        search_frame.grid(row=0, column=11, sticky="e", padx=(20, 0))
        
        search_label = tk.Label(
            search_frame, 
            text="Buscar:", 
            font=("Arial", 10),
            fg=PRIMARY_COLOR,
            bg=WHITE_COLOR
        )
        search_label.pack(side=tk.LEFT, padx=(0, 5))
        
        search_entry = tk.Entry(
            search_frame,
            textvariable=self.search_var,
            width=15,
            font=("Arial", 10),
            relief=tk.SOLID,
            borderwidth=1,
            fg=BLACK_COLOR  # Texto em preto conforme solicitado
        )
        search_entry.pack(side=tk.LEFT)
        
        # ===== PAINEL DE EDIÇÃO =====
        edit_frame = tk.Frame(main_container, bg=LIGHT_GRAY, padx=15, pady=15, relief=tk.GROOVE, borderwidth=1)
        edit_frame.grid(row=1, column=0, sticky="ew", pady=(0, 15))
        
        # Título do painel
        edit_title = tk.Label(
            edit_frame, 
            text="Editar Valor", 
            font=("Arial", 12, "bold"),
            fg=PRIMARY_COLOR,
            bg=LIGHT_GRAY
        )
        edit_title.grid(row=0, column=0, sticky="w", columnspan=6, pady=(0, 10))
        
        # Seleção de origem
        origem_label = tk.Label(
            edit_frame, 
            text="Origem:", 
            font=("Arial", 10, "bold"),
            fg=PRIMARY_COLOR,
            bg=LIGHT_GRAY
        )
        origem_label.grid(row=1, column=0, sticky="w", padx=(0, 5))
        
        self.origem_var = tk.StringVar()
        self.origem_combobox = ttk.Combobox(
            edit_frame, 
            values=self.estados,
            textvariable=self.origem_var,
            width=8,
            style="TCombobox"
        )
        self.origem_combobox.grid(row=1, column=1, padx=(0, 15))
        self.origem_combobox.set(self.estados[0])
        self.origem_combobox.bind("<<ComboboxSelected>>", self.update_selection)
        
        # Seleção de destino
        destino_label = tk.Label(
            edit_frame, 
            text="Destino:", 
            font=("Arial", 10, "bold"),
            fg=PRIMARY_COLOR,
            bg=LIGHT_GRAY
        )
        destino_label.grid(row=1, column=2, sticky="w", padx=(0, 5))
        
        self.destino_var = tk.StringVar()
        self.destino_combobox = ttk.Combobox(
            edit_frame, 
            values=self.estados,
            textvariable=self.destino_var,
            width=8,
            style="TCombobox"
        )
        self.destino_combobox.grid(row=1, column=3, padx=(0, 15))
        self.destino_combobox.set(self.estados[0])
        self.destino_combobox.bind("<<ComboboxSelected>>", self.update_selection)
        
        # Campo de valor
        valor_label = tk.Label(
            edit_frame, 
            text="Valor:", 
            font=("Arial", 10, "bold"),
            fg=PRIMARY_COLOR,
            bg=LIGHT_GRAY
        )
        valor_label.grid(row=1, column=4, sticky="w", padx=(0, 5))
        
        self.valor_var = tk.StringVar()
        self.valor_entry = tk.Entry(
            edit_frame,
            textvariable=self.valor_var,
            width=15,
            font=("Arial", 10),
            relief=tk.SOLID,
            borderwidth=1,
            fg=BLACK_COLOR  # Texto em preto conforme solicitado
        )
        self.valor_entry.grid(row=1, column=5, padx=(0, 15))
        self.valor_entry.bind("<Return>", self.set_value)
        
        # Botão de definir valor
        set_button = tk.Button(
            edit_frame, 
            text="Definir Valor", 
            command=self.set_value,
            bg=SECONDARY_COLOR,
            fg=WHITE_COLOR,
            activebackground=PRIMARY_COLOR,
            activeforeground=WHITE_COLOR,
            font=("Arial", 10),
            borderwidth=0,
            padx=15,
            pady=5,
            cursor="hand2"
        )
        set_button.grid(row=1, column=6, padx=(0, 5))
        
        # ===== TABELA DE DADOS =====
        table_container = tk.Frame(main_container, bg=WHITE_COLOR)
        table_container.grid(row=2, column=0, sticky="nsew")
        table_container.grid_columnconfigure(0, weight=1)
        table_container.grid_rowconfigure(0, weight=1)
        
        # Criar um frame com barras de rolagem
        table_frame = tk.Frame(table_container, bg=WHITE_COLOR)
        table_frame.grid(row=0, column=0, sticky="nsew")
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)
        
        # Criar canvas para permitir rolagem
        self.canvas = tk.Canvas(table_frame, bg=WHITE_COLOR, highlightthickness=0)
        scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.canvas.yview, style="TScrollbar")
        scrollbar_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.canvas.xview, style="TScrollbar")
        
        # Configurar o canvas
        self.canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # Posicionar scrollbars e canvas
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        self.canvas.grid(row=0, column=0, sticky="nsew")
        
        # Frame dentro do canvas para conter a tabela
        self.table_inner_frame = tk.Frame(self.canvas, bg=WHITE_COLOR)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.table_inner_frame, anchor="nw")
        
        # Configurar eventos para redimensionamento
        self.table_inner_frame.bind("<Configure>", self.on_frame_configure)
        self.canvas.bind("<Configure>", self.on_canvas_configure)
        
        # Criar a tabela de estados
        self.create_table()
        
        # ===== BARRA DE STATUS =====
        self.status_var = tk.StringVar()
        self.status_var.set("Pronto para uso")
        
        status_bar = tk.Label(
            self.root, 
            textvariable=self.status_var,
            fg=PRIMARY_COLOR,
            bg=LIGHT_GRAY,
            anchor="w",
            padx=15,
            pady=5,
            font=("Arial", 9)
        )
        status_bar.grid(row=2, column=0, sticky="ew")
        
        # Atualizar a seleção inicial
        self.update_selection()
    
    def on_frame_configure(self, event):
        # Atualizar a região de rolagem para acomodar o frame interno
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
    
    def on_canvas_configure(self, event):
        # Redimensionar a janela do canvas para preencher o espaço disponível
        width = event.width
        self.canvas.itemconfig(self.canvas_window, width=width)
    
    def create_table(self):
        # Limpar a tabela existente
        for widget in self.table_inner_frame.winfo_children():
            widget.destroy()
        
        # Configurar o grid
        for i in range(len(self.estados) + 1):
            self.table_inner_frame.grid_columnconfigure(i, weight=1, minsize=60)
            self.table_inner_frame.grid_rowconfigure(i, weight=1, minsize=30)
        
        # Criar cabeçalhos
        header_style = {
            "font": ("Arial", 10, "bold"),
            "bg": PRIMARY_COLOR,
            "fg": WHITE_COLOR,
            "relief": tk.RAISED,
            "padx": 5,
            "pady": 5,
            "width": 5
        }
        
        # Célula vazia no canto superior esquerdo
        tk.Label(
            self.table_inner_frame, 
            text="", 
            **header_style
        ).grid(row=0, column=0, sticky="nsew")
        
        # Cabeçalhos de coluna (estados de destino)
        for col, estado in enumerate(self.estados, 1):
            tk.Label(
                self.table_inner_frame, 
                text=estado, 
                **header_style
            ).grid(row=0, column=col, sticky="nsew")
        
        # Cabeçalhos de linha (estados de origem)
        for row, origem in enumerate(self.estados, 1):
            tk.Label(
                self.table_inner_frame, 
                text=origem, 
                **header_style
            ).grid(row=row, column=0, sticky="nsew")
            
            # Células de dados
            for col, destino in enumerate(self.estados, 1):
                key = f"{origem}-{destino}"
                
                # Determinar a cor de fundo da célula
                bg_color = LIGHT_GRAY if (row + col) % 2 == 0 else WHITE_COLOR
                
                # Criar um frame para a célula
                cell_frame = tk.Frame(
                    self.table_inner_frame, 
                    bg=bg_color,
                    relief=tk.SOLID,
                    borderwidth=1
                )
                cell_frame.grid(row=row, column=col, sticky="nsew", padx=1, pady=1)
                cell_frame.grid_columnconfigure(0, weight=1)
                cell_frame.grid_rowconfigure(0, weight=1)
                
                # Label para mostrar o valor
                valor = self.valores.get(key, "")
                cell_label = tk.Label(
                    cell_frame,
                    text=valor if valor else "",
                    bg=bg_color,
                    fg=BLACK_COLOR,  # Texto em preto conforme solicitado
                    padx=5,
                    pady=2,
                    anchor="center",
                    font=("Arial", 9)
                )
                cell_label.grid(row=0, column=0, sticky="nsew")
                
                # Configurar evento de clique para editar a célula
                cell_label.bind("<Button-1>", partial(self.edit_cell, origem, destino))
                
                # Armazenar referência ao label para atualização posterior
                self.valores[f"{origem}-{destino}_label"] = cell_label
    
    def edit_cell(self, origem, destino, event=None):
        # Atualizar os comboboxes
        self.origem_var.set(origem)
        self.destino_var.set(destino)
        
        # Atualizar o valor atual
        key = f"{origem}-{destino}"
        self.valor_var.set(self.valores.get(key, ""))
        
        # Armazenar a célula atual sendo editada
        self.current_edit = key
        
        # Destacar a célula sendo editada
        self.highlight_cell(origem, destino)
        
        # Focar no campo de entrada
        self.valor_entry.focus_set()
        self.valor_entry.select_range(0, tk.END)
        
        # Atualizar status
        self.status_var.set(f"Editando: {origem}-{destino}")
    
    def highlight_cell(self, origem, destino):
        # Remover destaque de todas as células
        for o in self.estados:
            for d in self.estados:
                key = f"{o}-{d}_label"
                if key in self.valores:
                    label = self.valores[key]
                    row = self.estados.index(o) + 1
                    col = self.estados.index(d) + 1
                    bg_color = LIGHT_GRAY if (row + col) % 2 == 0 else WHITE_COLOR
                    label.configure(bg=bg_color, fg=BLACK_COLOR)
        
        # Destacar a célula selecionada
        key = f"{origem}-{destino}_label"
        if key in self.valores:
            self.valores[key].configure(bg=SECONDARY_COLOR, fg=WHITE_COLOR)
    
    def update_selection(self, event=None):
        origem = self.origem_var.get()
        destino = self.destino_var.get()
        key = f"{origem}-{destino}"
        
        # Atualizar o valor no campo de entrada
        self.valor_var.set(self.valores.get(key, ""))
        
        # Armazenar a célula atual sendo editada
        self.current_edit = key
        
        # Destacar a célula
        self.highlight_cell(origem, destino)
        
        # Atualizar status
        self.status_var.set(f"Selecionado: {origem}-{destino}")
    
    def set_value(self, event=None):
        if not self.current_edit:
            return
        
        # Obter o valor do campo de entrada
        valor = self.valor_var.get()
        
        # Atualizar o valor no dicionário
        self.valores[self.current_edit] = valor
        
        # Atualizar o label na tabela
        key = f"{self.current_edit}_label"
        if key in self.valores:
            self.valores[key].configure(text=valor if valor else "")
        
        # Atualizar status
        self.status_var.set(f"Valor atualizado: {self.current_edit} = {valor}")
        
        # Mover para a próxima célula
        self.move_to_next_cell()
    
    def move_to_next_cell(self):
        if not self.current_edit:
            return
        
        # Obter origem e destino atuais
        origem, destino = self.current_edit.split("-")
        
        # Encontrar o próximo estado de destino
        idx_destino = self.estados.index(destino)
        if idx_destino < len(self.estados) - 1:
            # Mover para o próximo estado de destino na mesma linha
            novo_destino = self.estados[idx_destino + 1]
            self.destino_var.set(novo_destino)
        else:
            # Mover para o primeiro estado de destino na próxima linha
            idx_origem = self.estados.index(origem)
            if idx_origem < len(self.estados) - 1:
                novo_origem = self.estados[idx_origem + 1]
                self.origem_var.set(novo_origem)
                self.destino_var.set(self.estados[0])
        
        # Atualizar a seleção
        self.update_selection()
        
        # Focar no campo de entrada
        self.valor_entry.focus_set()
        self.valor_entry.select_range(0, tk.END)
    
    def filter_table(self, *args):
        search_text = self.search_var.get().lower()
        
        if not search_text:
            # Se a busca estiver vazia, mostrar todas as células
            for origem in self.estados:
                for destino in self.estados:
                    key = f"{origem}-{destino}_label"
                    if key in self.valores:
                        self.valores[key].master.grid()
            return
        
        # Filtrar células com base no texto de busca
        for origem in self.estados:
            for destino in self.estados:
                key = f"{origem}-{destino}"
                label_key = f"{key}_label"
                
                if label_key in self.valores:
                    valor = self.valores.get(key, "").lower()
                    
                    # Verificar se o texto de busca está no valor, origem ou destino
                    if (search_text in valor or 
                        search_text in origem.lower() or 
                        search_text in destino.lower()):
                        # Mostrar a célula
                        self.valores[label_key].master.grid()
                    else:
                        # Ocultar a célula
                        self.valores[label_key].master.grid_remove()
    
    def save_data(self):
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                initialfile="estados_valores.json"
            )
            
            if not file_path:  # Se o usuário cancelar
                return
            
            # Criar um dicionário limpo sem as referências aos labels
            clean_data = {}
            for key, value in self.valores.items():
                if not key.endswith("_label"):
                    clean_data[key] = value
                
            with open(file_path, 'w') as f:
                json.dump(clean_data, f)
            
            self.status_var.set(f"Dados salvos em {file_path}")
            messagebox.showinfo("Sucesso", f"Dados salvos com sucesso em {file_path}")
        except Exception as e:
            self.status_var.set(f"Erro ao salvar dados: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao salvar dados: {str(e)}")
    
    def load_data(self):
        try:
            file_path = filedialog.askopenfilename(
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )
            
            if not file_path:  # Se o usuário cancelar
                return
                
            if os.path.exists(file_path):
                with open(file_path, 'r') as f:
                    loaded_data = json.load(f)
                
                # Atualizar apenas as chaves válidas (combinações de estados)
                for key, value in loaded_data.items():
                    if "-" in key and not key.endswith("_label"):
                        self.valores[key] = value
                
                # Recriar a tabela com os novos valores
                self.create_table()
                
                # Atualizar a seleção atual
                self.update_selection()
                
                self.status_var.set(f"Dados carregados de {file_path}")
                messagebox.showinfo("Sucesso", f"Dados carregados com sucesso de {file_path}")
            else:
                self.status_var.set(f"Arquivo {file_path} não encontrado")
                messagebox.showwarning("Aviso", f"Arquivo {file_path} não encontrado")
        except Exception as e:
            self.status_var.set(f"Erro ao carregar dados: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao carregar dados: {str(e)}")
    
    def export_to_csv(self):
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile="estados_export.csv"
            )
            
            if not file_path:  # Se o usuário cancelar
                return
            
            # Criar um DataFrame pandas com os dados
            data = []
            for origem in self.estados:
                row = [origem]
                for destino in self.estados:
                    key = f"{origem}-{destino}"
                    row.append(self.valores.get(key, ""))
                data.append(row)
            
            # Criar o DataFrame
            columns = ["Estado"] + self.estados
            df = pd.DataFrame(data, columns=columns)
            
            # Salvar como CSV
            df.to_csv(file_path, index=False)
            
            self.status_var.set(f"Dados exportados para {file_path}")
            messagebox.showinfo("Sucesso", f"Dados exportados com sucesso para {file_path}")
        except Exception as e:
            self.status_var.set(f"Erro ao exportar dados: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao exportar dados: {str(e)}")
    
    def export_to_excel(self):
        try:
            # Verificar se o pandas tem suporte a Excel
            if not hasattr(pd.DataFrame, 'to_excel'):
                messagebox.showwarning(
                    "Dependência Faltando", 
                    "A exportação para Excel requer a biblioteca openpyxl.\n"
                    "Instale-a usando: pip install openpyxl"
                )
                return
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="estados_export.xlsx"
            )
            
            if not file_path:  # Se o usuário cancelar
                return
            
            # Criar um DataFrame pandas com os dados
            data = []
            for origem in self.estados:
                row = [origem]
                for destino in self.estados:
                    key = f"{origem}-{destino}"
                    row.append(self.valores.get(key, ""))
                data.append(row)
            
            # Criar o DataFrame
            columns = ["Estado"] + self.estados
            df = pd.DataFrame(data, columns=columns)
            
            # Salvar como Excel
            df.to_excel(file_path, index=False, sheet_name="Combinações de Estados")
            
            # Formatar o Excel (opcional)
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # Abrir o arquivo Excel
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                # Definir cores
                primary_color_fill = PatternFill(start_color="56239F", end_color="56239F", fill_type="solid")
                secondary_color_fill = PatternFill(start_color="FF7200", end_color="FF7200", fill_type="solid")
                light_gray_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
                
                # Definir fontes
                white_font = Font(color="FFFFFF", bold=True)
                black_font = Font(color="000000")
                
                # Definir bordas
                thin_border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                
                # Formatar cabeçalhos
                for col in range(1, len(self.estados) + 2):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = primary_color_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                for row in range(2, len(self.estados) + 2):
                    cell = ws.cell(row=row, column=1)
                    cell.fill = primary_color_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                # Formatar células de dados
                for row in range(2, len(self.estados) + 2):
                    for col in range(2, len(self.estados) + 2):
                        cell = ws.cell(row=row, column=col)
                        if (row + col) % 2 == 0:
                            cell.fill = light_gray_fill
                        cell.font = black_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                
                # Ajustar largura das colunas
                for col in range(1, len(self.estados) + 2):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
                
                # Salvar as alterações
                wb.save(file_path)
            except ImportError:
                # Se openpyxl não estiver disponível para formatação avançada, apenas continuar
                pass
            
            self.status_var.set(f"Dados exportados para {file_path}")
            messagebox.showinfo("Sucesso", f"Dados exportados com sucesso para {file_path}")
        except Exception as e:
            self.status_var.set(f"Erro ao exportar dados: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao exportar dados: {str(e)}")
    
    def import_from_csv(self):
        try:
            file_path = filedialog.askopenfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            
            if not file_path:  # Se o usuário cancelar
                return
            
            # Ler o arquivo CSV
            df = pd.read_csv(file_path)
            
            # Verificar se o formato é compatível
            if len(df.columns) < 2 or df.columns[0] != "Estado":
                messagebox.showerror(
                    "Formato Inválido", 
                    "O arquivo CSV deve ter uma coluna 'Estado' seguida por colunas para cada estado."
                )
                return
            
            # Extrair os estados das colunas
            csv_estados = df.columns[1:].tolist()
            
            # Verificar se os estados são compatíveis
            if not all(estado in self.estados for estado in csv_estados):
                messagebox.showwarning(
                    "Estados Incompatíveis", 
                    "Alguns estados no arquivo CSV não correspondem aos estados do aplicativo.\n"
                    "Os dados serão importados apenas para os estados correspondentes."
                )
            
            # Limpar os valores atuais
            for origem in self.estados:
                for destino in self.estados:
                    self.valores[f"{origem}-{destino}"] = ""
            
            # Importar os dados
            for _, row in df.iterrows():
                origem = row["Estado"]
                if origem in self.estados:
                    for destino in csv_estados:
                        if destino in self.estados:
                            valor = str(row[destino])
                            # Tratar valores NaN
                            if valor.lower() == "nan":
                                valor = ""
                            self.valores[f"{origem}-{destino}"] = valor
            
            # Recriar a tabela com os novos valores
            self.create_table()
            
            # Atualizar a seleção atual
            self.update_selection()
            
            self.status_var.set(f"Dados importados de {file_path}")
            messagebox.showinfo("Sucesso", f"Dados importados com sucesso de {file_path}")
        except Exception as e:
            self.status_var.set(f"Erro ao importar dados: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao importar dados: {str(e)}")
    
    def import_from_excel(self):
        try:
            # Verificar se o pandas tem suporte a Excel
            if not hasattr(pd.DataFrame, 'read_excel'):
                messagebox.showwarning(
                    "Dependência Faltando", 
                    "A importação de Excel requer a biblioteca openpyxl.\n"
                    "Instale-a usando: pip install openpyxl"
                )
                return
            
            file_path = filedialog.askopenfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("Excel files", "*.xls"), ("All files", "*.*")]
            )
            
            if not file_path:  # Se o usuário cancelar
                return
            
            # Ler o arquivo Excel
            df = pd.read_excel(file_path)
            
            # Verificar se o formato é compatível
            if len(df.columns) < 2 or df.columns[0] != "Estado":
                messagebox.showerror(
                    "Formato Inválido", 
                    "O arquivo Excel deve ter uma coluna 'Estado' seguida por colunas para cada estado."
                )
                return
            
            # Extrair os estados das colunas
            excel_estados = df.columns[1:].tolist()
            
            # Verificar se os estados são compatíveis
            if not all(estado in self.estados for estado in excel_estados):
                messagebox.showwarning(
                    "Estados Incompatíveis", 
                    "Alguns estados no arquivo Excel não correspondem aos estados do aplicativo.\n"
                    "Os dados serão importados apenas para os estados correspondentes."
                )
            
            # Limpar os valores atuais
            for origem in self.estados:
                for destino in self.estados:
                    self.valores[f"{origem}-{destino}"] = ""
            
            # Importar os dados
            for _, row in df.iterrows():
                origem = row["Estado"]
                if origem in self.estados:
                    for destino in excel_estados:
                        if destino in self.estados:
                            valor = str(row[destino])
                            # Tratar valores NaN
                            if valor.lower() == "nan":
                                valor = ""
                            self.valores[f"{origem}-{destino}"] = valor
            
            # Recriar a tabela com os novos valores
            self.create_table()
            
            # Atualizar a seleção atual
            self.update_selection()
            
            self.status_var.set(f"Dados importados de {file_path}")
            messagebox.showinfo("Sucesso", f"Dados importados com sucesso de {file_path}")
        except Exception as e:
            self.status_var.set(f"Erro ao importar dados: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao importar dados: {str(e)}")

def main():
    root = tk.Tk()
    app = EstadosApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()